import openpyxl

workbook1 = openpyxl.Workbook()

print(workbook1.active.title)
print(workbook1.sheetnames)

workbook1['Sheet'].title = "CartDetails"
worksheet1 = workbook1.active
worksheet1.cell(1, 1).value = 'Product ID'
worksheet1['A1']
worksheet1['B1'].value = 'Product Name'
worksheet1['C1'].value = 'Product Price'
worksheet1['A2'].value = 1
worksheet1['B2'].value = 'One Plus 11R'
worksheet1['C2'].value = 40000.00
worksheet1['A3'].value = 2
worksheet1['B3'].value = 'Nothing Phone 1'
worksheet1['C3'].value = 39999.00
worksheet1['A4'].value = 3
worksheet1['B4'].value = 'Apple iPhone 14'
worksheet1['C4'].value = 145000.00

workbook1.save("Phone1.xlsx")

# print(worksheet1.calculate_dimension())
# print(worksheet1.path)
# print(worksheet1.views)
# print(worksheet1.values)
# print(worksheet1.tables)
# print(worksheet1.sheet_view)
# print(worksheet1.print_area())
