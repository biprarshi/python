#copy paste excel files
import openpyxl

file1 = 'test.xlsx'
workbook1 = openpyxl.load_workbook(file1)
worksheet1 = workbook1.worksheets[1]


file2 = 'test_copy.xlsx'
workbook2 = openpyxl.Workbook()
workbook2.active.title = 'copied_sheet'
workbook2.save(file2)
worksheet2 = workbook2.worksheets[0]


max_r1 = worksheet1.max_row
max_c1 = worksheet1.max_column

for row1 in range(1, max_r1+1):
    for column1 in range(1,max_c1+1):
        value1 = worksheet1.cell(row1,column1).value
        worksheet2.cell(row=row1,column=column1).value = value1

workbook2.save(file2)
