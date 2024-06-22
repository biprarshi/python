# continuously delete row 2 until there is only a single row left over
import openpyxl


def delete_row2(sheet1):
    while sheet1.max_row > 1:
        sheet1.delete_rows(2, 1)
    return


if __name__ == "__main__":
    path1 = 'test.xlsx'
    workbook1 = openpyxl.load_workbook(path1)
    worksheet1 = workbook1['wrksheet3']
    print("Maximum rows before deleting:", worksheet1.max_row)
    delete_row2(worksheet1)
    print("Maximum rows after deleting:", worksheet1.max_row)
    path1 = "delete_row2.xlsx"
    workbook1.save(path1)

wrkbk1 = openpyxl.Workbook()
# wrkbk1.active.title = "tomato"
wrkbk1.save("wrkbk1.xlsx")
wrksht1 = wrkbk1.worksheets[0]
# wrksht1.title = "tomato2"
wrksht1.cell(row=1, column=1).value = "Values"
for i in range(50):
    wrksht1.cell(row=i+2, column=1).value = i+1
wrkbk1.save("wrkbk1.xlsx")

wrksht1.delete_rows(1, 0)
wrkbk1.save("wrkbk1.xlsx")
