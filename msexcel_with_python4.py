# function to remove empty rows

import openpyxl

'''
def remove_empty_rows(sheet1, row1):
    for cell1 in row1:
        # print(cell1, cell1.value, sep='@@')
        # print(row1, row1[0].row, row1[0].value, sep='@@')
        if cell1.value != None:
            return
        else:
            # get the row number from the first cell and remove the row
            # sheet1.delete_rows(row1[0].row, 1)
            print(row1[0].row)
            for x in sheet1:
                print(x[0].row,end='. ')
                for y in x:
                    print(y.value, end=' ')
                print()
            sheet1.delete_rows(4, 1)
            for x in sheet1:
                print(x[0].row, end='. ')
                for y in x:
                    print(y.value, end=' ')
                print()
        print("="*10)
'''


def remove_empty_rows(sheet1, row1):
    for cell1 in row1:
        if cell1.value == None:
            sheet1.delete_rows(row1[0].row, 1)
        else:
            return

if __name__ == '__main__':
    path1 = 'test.xlsx'
    workbook1 = openpyxl.load_workbook(path1)
    worksheet1 = workbook1['wrksheet5']
    print("Maximum rows before deleting:", worksheet1.max_row)
    for rows2 in worksheet1:
        # print(rows2)
        # for x in rows2:
        # print(x)
        remove_empty_rows(worksheet1, rows2)

        # worksheet1.delete_rows(4,2)
    print("Maximum rows after deleting:", worksheet1.max_row)
    path1 = "empty_rows_deleted.xlsx"
    workbook1.save(path1)

    # wb1 = openpyxl.load_workbook(path1)
    # ws1 = wb1['wrksheet5']
    # print(ws1.max_row)
