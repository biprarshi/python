import openpyxl

workbook1 = openpyxl.load_workbook("charts_msexcel1.xlsx")

sheets1 = workbook1.sheetnames
print(sheets1)
print(workbook1.active.title)

worksheet1 = workbook1['wrksheet1']
value1 = worksheet1['A3'].value
print(value1)
value1 = workbook1['wrksheet1']['A3'].value
print(value1)
value1 = workbook1['wrksheet1'].cell(3, 1).value
print(value1)
value1 = workbook1['wrksheet1'].cell(column=1, row=3).value
print(value1)

# Number of row & column:
rows1 = worksheet1.max_row
columns1 = worksheet1.max_column
print(f"rows:", rows1)
print(f'columns:', columns1)

# Read the value of all row and coulumns:
for i in range(1, rows1+1):
    for j in range(1, columns1+1):
        print(worksheet1.cell(i, j).value)

worksheet1 = workbook1['wrksheet5']
worksheet1.cell(row=9, column=1).value = 'Amit'
worksheet1.cell(row=9, column=2).value = 'SaltLake'
worksheet1.cell(row=9, column=3).value = 5
worksheet1.cell(row=10, column=1, value='Ankana')
worksheet1.cell(row=10, column=2, value='Sector3')
worksheet1.cell(row=10, column=3, value=10)
workbook1.save('charts_msexcel1.xlsx')
print("Data inserted succeessfully")

# To save older datials with new one in a new File:
# if file is not present then it created by it self.
worksheet1.cell(row=11, column=1).value = 'Raju'
worksheet1.cell(row=11, column=2).value = 'OMAN'
worksheet1.cell(row=11, column=3,).value = 19
workbook1.save("charts_msexcel1_updated.xlsx")
print("Data updated successfully")
# workbook1.create_sheet(title="papa")
# workbook1.save("papa.xlsx")